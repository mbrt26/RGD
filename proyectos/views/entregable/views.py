from django.views.generic import ListView, CreateView, UpdateView, DetailView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.db.models import Q, Count
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Case, When, Value, IntegerField
from datetime import timedelta
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_http_methods
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
import io
from proyectos.models import EntregaDocumental, Proyecto, EntregableProyecto
from proyectos.forms.entregable_forms import (
    EntregableProyectoForm, ConfiguracionMasivaForm, FiltroEntregablesForm, 
    EntregablePersonalizadoForm, EntregableImportForm
)


class EntregaDocumentalListView(LoginRequiredMixin, ListView):
    model = EntregaDocumental
    template_name = 'proyectos/entregable/list.html'
    context_object_name = 'entregables'
    ordering = ['-fecha_entrega']
    title = 'Entregables'
    add_url = 'proyectos:entregable_create'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = self.title
        context['view'] = {'add_url': self.add_url}
        return context


class EntregaDocumentalCreateView(LoginRequiredMixin, CreateView):
    model = EntregaDocumental
    template_name = 'proyectos/entregable/form.html'
    fields = ['nombre', 'proyecto', 'descripcion', 'fecha_entrega', 'estado', 'archivo']
    success_url = reverse_lazy('proyectos:entregable_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Nuevo Entregable'
        return context


class EntregaDocumentalDetailView(LoginRequiredMixin, DetailView):
    model = EntregaDocumental
    template_name = 'proyectos/entregable/detail.html'
    context_object_name = 'entregable'


class EntregaDocumentalUpdateView(LoginRequiredMixin, UpdateView):
    model = EntregaDocumental
    template_name = 'proyectos/entregable/form.html'
    fields = ['nombre', 'proyecto', 'descripcion', 'fecha_entrega', 'estado', 'archivo']
    success_url = reverse_lazy('proyectos:entregable_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Editar Entregable: {self.object}'
        return context


class GestionEntregablesView(LoginRequiredMixin, TemplateView):
    """Vista principal para gestionar los entregables de un proyecto"""
    template_name = 'proyectos/entregables/gestion.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proyectos'] = Proyecto.objects.all().order_by('-fecha_creacion')
        
        # Si se seleccionó un proyecto
        proyecto_id = self.request.GET.get('proyecto')
        if proyecto_id:
            try:
                proyecto = Proyecto.objects.get(id=proyecto_id)
                context['proyecto_seleccionado'] = proyecto
                
                # Cargar entregables desde JSON si no existen
                if not proyecto.entregables_proyecto.exists():
                    EntregableProyecto.cargar_entregables_desde_json(proyecto)
                
                # Obtener entregables agrupados por fase
                entregables = proyecto.entregables_proyecto.all()
                context['entregables_por_fase'] = {
                    'Definición': entregables.filter(fase='Definición'),
                    'Planeación': entregables.filter(fase='Planeación'),
                    'Ejecución': entregables.filter(fase='Ejecución'),
                    'Entrega': entregables.filter(fase='Entrega'),
                }
                
            except Proyecto.DoesNotExist:
                messages.error(self.request, 'Proyecto no encontrado.')
        
        return context
    
    def post(self, request, *args, **kwargs):
        """Procesar la actualización de entregables seleccionados"""
        proyecto_id = request.POST.get('proyecto_id')
        if not proyecto_id:
            messages.error(request, 'Debe seleccionar un proyecto.')
            return redirect('proyectos:gestion_entregables')
        
        try:
            proyecto = Proyecto.objects.get(id=proyecto_id)
            
            # Actualizar selección de entregables
            entregables_seleccionados = request.POST.getlist('entregables')
            
            # Primero, deseleccionar todos los no obligatorios
            proyecto.entregables_proyecto.filter(obligatorio=False).update(seleccionado=False)
            
            # Luego, seleccionar los marcados
            for entregable_id in entregables_seleccionados:
                try:
                    entregable = proyecto.entregables_proyecto.get(id=entregable_id)
                    entregable.seleccionado = True
                    entregable.save()
                except EntregableProyecto.DoesNotExist:
                    continue
            
            messages.success(request, f'Entregables actualizados para el proyecto {proyecto.nombre_proyecto}')
            return redirect(f'{reverse_lazy("proyectos:gestion_entregables")}?proyecto={proyecto_id}')
            
        except Proyecto.DoesNotExist:
            messages.error(request, 'Proyecto no encontrado.')
            return redirect('proyectos:gestion_entregables')


class EntregableProyectoUpdateView(LoginRequiredMixin, UpdateView):
    """Vista para actualizar un entregable específico del proyecto"""
    model = EntregableProyecto
    template_name = 'proyectos/entregables/entregable_form.html'
    fields = ['estado', 'archivo', 'fecha_entrega', 'observaciones']
    
    def get_success_url(self):
        return f'{reverse_lazy("proyectos:gestion_entregables")}?proyecto={self.object.proyecto.id}'
    
    def form_valid(self, form):
        messages.success(self.request, f'Entregable {self.object.codigo} actualizado correctamente.')
        return super().form_valid(form)


@require_http_methods(["POST"])
def entregable_proyecto_update_inline(request, pk):
    """Vista AJAX para actualizar entregable inline"""
    try:
        entregable = get_object_or_404(EntregableProyecto, pk=pk)
        
        # Obtener datos del request
        fecha_entrega = request.POST.get('fecha_entrega')
        estado = request.POST.get('estado')
        archivo = request.FILES.get('archivo')
        
        # Validar y actualizar campos
        if fecha_entrega:
            from datetime import datetime
            try:
                entregable.fecha_entrega = datetime.strptime(fecha_entrega, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'errors': {'fecha_entrega': 'Formato de fecha inválido'}
                })
        
        if estado and estado in ['pendiente', 'en_proceso', 'completado', 'no_aplica']:
            entregable.estado = estado
        
        if archivo:
            entregable.archivo = archivo
        
        # Guardar cambios
        entregable.save()
        
        # Preparar respuesta
        response_data = {
            'success': True,
            'message': 'Entregable actualizado correctamente',
            'entregable': {
                'id': entregable.id,
                'fecha_entrega': entregable.fecha_entrega.strftime('%Y-%m-%d') if entregable.fecha_entrega else None,
                'estado': entregable.estado,
                'estado_display': entregable.get_estado_display(),
            }
        }
        
        if entregable.archivo:
            response_data['archivo_url'] = entregable.archivo.url
        
        return JsonResponse(response_data)
        
    except EntregableProyecto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Entregable no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


def cargar_entregables_proyecto(request, proyecto_id):
    """Vista AJAX para cargar entregables de un proyecto"""
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
        
        # Cargar entregables desde JSON si no existen
        if not proyecto.entregables_proyecto.exists():
            EntregableProyecto.cargar_entregables_desde_json(proyecto)
        
        entregables_data = []
        for entregable in proyecto.entregables_proyecto.all():
            entregables_data.append({
                'id': entregable.id,
                'codigo': entregable.codigo,
                'nombre': entregable.nombre,
                'fase': entregable.fase,
                'obligatorio': entregable.obligatorio,
                'seleccionado': entregable.seleccionado,
                'estado': entregable.estado,
                'creador': entregable.creador,
                'consolidador': entregable.consolidador,
            })
        
        return JsonResponse({
            'success': True,
            'entregables': entregables_data,
            'proyecto': {
                'id': proyecto.id,
                'nombre': proyecto.nombre_proyecto,
                'cliente': proyecto.cliente
            }
        })
        
    except Proyecto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Proyecto no encontrado'
        }, status=404)


class EntregablesDashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard principal con estadísticas de entregables"""
    template_name = 'proyectos/entregables/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Estadísticas generales
        context['stats'] = {
            'proyectos_con_entregables': Proyecto.objects.filter(
                entregables_proyecto__isnull=False
            ).distinct().count(),
            'entregables_pendientes': EntregableProyecto.objects.filter(
                estado='pendiente', seleccionado=True
            ).count(),
            'entregables_completados': EntregableProyecto.objects.filter(
                estado='completado'
            ).count(),
            'entregables_vencidos': EntregableProyecto.objects.filter(
                fecha_entrega__lt=timezone.now().date(),
                estado__in=['pendiente', 'en_proceso']
            ).count()
        }
        
        # Entregables próximos a vencer (próximos 7 días)
        fecha_limite = timezone.now().date() + timedelta(days=7)
        context['entregables_proximos'] = EntregableProyecto.objects.filter(
            fecha_entrega__lte=fecha_limite,
            fecha_entrega__gte=timezone.now().date(),
            estado__in=['pendiente', 'en_proceso'],
            seleccionado=True
        ).select_related('proyecto')[:10]
        
        # Proyectos sin entregables configurados
        context['proyectos_sin_entregables'] = Proyecto.objects.filter(
            entregables_proyecto__isnull=True,
            estado__in=['en_progreso', 'iniciado']
        )[:5]
        
        return context


class ConfiguracionMasivaEntregablesView(LoginRequiredMixin, TemplateView):
    """Vista para configurar entregables en múltiples proyectos"""
    template_name = 'proyectos/entregables/configuracion_masiva.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proyectos'] = Proyecto.objects.filter(
            estado__in=['en_progreso', 'iniciado']
        ).order_by('-fecha_creacion')
        return context
    
    def post(self, request, *args, **kwargs):
        proyectos_ids = request.POST.getlist('proyectos')
        accion = request.POST.get('accion')
        
        if accion == 'cargar_entregables':
            for proyecto_id in proyectos_ids:
                try:
                    proyecto = Proyecto.objects.get(id=proyecto_id)
                    if not proyecto.entregables_proyecto.exists():
                        EntregableProyecto.cargar_entregables_desde_json(proyecto)
                except Proyecto.DoesNotExist:
                    continue
                    
            messages.success(request, f'Entregables cargados para {len(proyectos_ids)} proyectos.')
        
        return redirect('proyectos:configuracion_masiva_entregables')


class EntregablesFiltradosView(LoginRequiredMixin, ListView):
    """Vista con filtros avanzados para entregables"""
    model = EntregableProyecto
    template_name = 'proyectos/entregables/filtrados.html'
    context_object_name = 'entregables'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = EntregableProyecto.objects.select_related('proyecto')
        
        # Filtros
        fase = self.request.GET.get('fase')
        estado = self.request.GET.get('estado')
        proyecto_id = self.request.GET.get('proyecto')
        obligatorio = self.request.GET.get('obligatorio')
        vencidos = self.request.GET.get('vencidos')
        
        if fase:
            queryset = queryset.filter(fase=fase)
        if estado:
            queryset = queryset.filter(estado=estado)
        if proyecto_id:
            queryset = queryset.filter(proyecto_id=proyecto_id)
        if obligatorio:
            queryset = queryset.filter(obligatorio=obligatorio == 'true')
        if vencidos == 'true':
            queryset = queryset.filter(
                fecha_entrega__lt=timezone.now().date(),
                estado__in=['pendiente', 'en_proceso']
            )
            
        return queryset.order_by('-fecha_actualizacion')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proyectos'] = Proyecto.objects.all().order_by('nombre_proyecto')
        context['fases'] = EntregableProyecto.PHASE_CHOICES
        context['estados'] = EntregableProyecto.ESTADO_CHOICES
        
        # Fechas para comparación en template
        context['today'] = timezone.now().date()
        context['fecha_limite'] = timezone.now().date() + timedelta(days=7)
        
        # Mantener filtros seleccionados
        context['filtros'] = {
            'fase': self.request.GET.get('fase', ''),
            'estado': self.request.GET.get('estado', ''),
            'proyecto': self.request.GET.get('proyecto', ''),
            'obligatorio': self.request.GET.get('obligatorio', ''),
            'vencidos': self.request.GET.get('vencidos', ''),
        }
        
        return context


class ReporteEntregablesView(LoginRequiredMixin, TemplateView):
    """Genera reportes de entregables en Excel"""
    
    def get(self, request, *args, **kwargs):
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Entregables"
        
        # Headers
        headers = [
            'Proyecto', 'Cliente', 'Código', 'Nombre', 'Fase', 'Estado',
            'Obligatorio', 'Seleccionado', 'Fecha Entrega', 'Creador',
            'Consolidador', 'Archivo'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Datos
        entregables = EntregableProyecto.objects.select_related('proyecto').all()
        for row, entregable in enumerate(entregables, 2):
            ws.cell(row=row, column=1, value=entregable.proyecto.nombre_proyecto)
            ws.cell(row=row, column=2, value=entregable.proyecto.cliente)
            ws.cell(row=row, column=3, value=entregable.codigo)
            ws.cell(row=row, column=4, value=entregable.nombre)
            ws.cell(row=row, column=5, value=entregable.fase)
            ws.cell(row=row, column=6, value=entregable.get_estado_display())
            ws.cell(row=row, column=7, value='Sí' if entregable.obligatorio else 'No')
            ws.cell(row=row, column=8, value='Sí' if entregable.seleccionado else 'No')
            ws.cell(row=row, column=9, value=entregable.fecha_entrega)
            ws.cell(row=row, column=10, value=entregable.creador)
            ws.cell(row=row, column=11, value=entregable.consolidador)
            ws.cell(row=row, column=12, value='Sí' if entregable.archivo else 'No')
        
        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=reporte_entregables.xlsx'
        
        # Guardar en memoria
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        
        return response


class EntregablePersonalizadoCreateView(LoginRequiredMixin, CreateView):
    """Vista para crear entregables personalizados"""
    model = EntregableProyecto
    form_class = EntregablePersonalizadoForm
    template_name = 'proyectos/entregables/entregable_personalizado_form.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.proyecto = get_object_or_404(Proyecto, id=kwargs['proyecto_id'])
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['proyecto'] = self.proyecto
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proyecto'] = self.proyecto
        context['title'] = f'Agregar Entregable Personalizado - {self.proyecto.nombre_proyecto}'
        return context
    
    def get_success_url(self):
        messages.success(self.request, f'Entregable personalizado creado exitosamente para el proyecto {self.proyecto.nombre_proyecto}')
        return f'{reverse_lazy("proyectos:gestion_entregables")}?proyecto={self.proyecto.id}'


class CambiarTipoEntregableView(LoginRequiredMixin, TemplateView):
    """Vista para cambiar el tipo de entregables (obligatorio/opcional)"""
    template_name = 'proyectos/entregables/cambiar_tipo.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.proyecto = get_object_or_404(Proyecto, id=kwargs['proyecto_id'])
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proyecto'] = self.proyecto
        
        # Obtener entregables agrupados por fase
        entregables = self.proyecto.entregables_proyecto.all()
        context['entregables_por_fase'] = {
            'Definición': entregables.filter(fase='Definición'),
            'Planeación': entregables.filter(fase='Planeación'),
            'Ejecución': entregables.filter(fase='Ejecución'),
            'Entrega': entregables.filter(fase='Entrega'),
        }
        
        return context
    
    def post(self, request, *args, **kwargs):
        """Procesar cambios de tipo de entregables"""
        entregables_obligatorios = request.POST.getlist('entregables_obligatorios')
        
        # Primero, marcar todos como opcionales
        self.proyecto.entregables_proyecto.update(obligatorio=False)
        
        # Luego, marcar los seleccionados como obligatorios
        if entregables_obligatorios:
            self.proyecto.entregables_proyecto.filter(
                id__in=entregables_obligatorios
            ).update(obligatorio=True, seleccionado=True)
        
        messages.success(request, f'Tipos de entregables actualizados para el proyecto {self.proyecto.nombre_proyecto}')
        return redirect(f'{reverse_lazy("proyectos:gestion_entregables")}?proyecto={self.proyecto.id}')


class EntregablesChecklistView(LoginRequiredMixin, TemplateView):
    """Vista de checklist para gestión visual de entregables"""
    template_name = 'proyectos/entregables/checklist.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.proyecto = get_object_or_404(Proyecto, id=kwargs['proyecto_id'])
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proyecto'] = self.proyecto
        
        # Cargar entregables si no existen
        if not self.proyecto.entregables_proyecto.exists():
            EntregableProyecto.cargar_entregables_desde_json(self.proyecto)
        
        # Obtener solo entregables seleccionados
        entregables = self.proyecto.entregables_proyecto.filter(seleccionado=True)
        
        context['entregables_por_fase'] = {
            'Definición': entregables.filter(fase='Definición'),
            'Planeación': entregables.filter(fase='Planeación'), 
            'Ejecución': entregables.filter(fase='Ejecución'),
            'Entrega': entregables.filter(fase='Entrega'),
        }
        
        # Estadísticas para el checklist
        total_entregables = entregables.count()
        completados = entregables.filter(estado='completado').count()
        en_proceso = entregables.filter(estado='en_proceso').count()
        pendientes = entregables.filter(estado='pendiente').count()
        no_aplica = entregables.filter(estado='no_aplica').count()
        
        context['stats'] = {
            'total': total_entregables,
            'completados': completados,
            'en_proceso': en_proceso,
            'pendientes': pendientes,
            'no_aplica': no_aplica,
            'porcentaje_completado': round((completados + no_aplica) / total_entregables * 100, 1) if total_entregables > 0 else 0,
        }
        
        # Entregables próximos a vencer (7 días)
        fecha_limite = timezone.now().date() + timedelta(days=7)
        context['proximos_vencer'] = entregables.filter(
            fecha_entrega__lte=fecha_limite,
            fecha_entrega__gte=timezone.now().date(),
            estado__in=['pendiente', 'en_proceso']
        ).order_by('fecha_entrega')
        
        return context


class EntregableImportView(LoginRequiredMixin, TemplateView):
    """Vista para importar entregables desde Excel"""
    template_name = 'proyectos/entregables/import.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = EntregableImportForm()
        return context
    
    def post(self, request, *args, **kwargs):
        form = EntregableImportForm(request.POST, request.FILES)
        
        if form.is_valid():
            try:
                import pandas as pd
                
                archivo_excel = form.cleaned_data['archivo_excel']
                proyecto = form.cleaned_data['proyecto']
                reemplazar_existentes = form.cleaned_data['reemplazar_existentes']
                
                # Leer el archivo Excel
                df = pd.read_excel(archivo_excel)
                
                # Validar columnas requeridas
                required_columns = ['codigo', 'nombre', 'fase', 'creador', 'consolidador']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    messages.error(
                        request, 
                        f'El archivo Excel debe contener las columnas: {", ".join(missing_columns)}'
                    )
                    return self.render_to_response({'form': form})
                
                # Procesar cada fila
                entregables_creados = 0
                entregables_actualizados = 0
                errores = []
                
                for index, row in df.iterrows():
                    try:
                        # Validar datos básicos
                        codigo = str(row['codigo']).strip()
                        nombre = str(row['nombre']).strip()
                        fase = str(row['fase']).strip()
                        creador = str(row['creador']).strip()
                        consolidador = str(row['consolidador']).strip()
                        
                        if not all([codigo, nombre, fase, creador, consolidador]):
                            errores.append(f'Fila {index + 2}: Campos obligatorios vacíos')
                            continue
                        
                        # Validar fase
                        fases_validas = [choice[0] for choice in EntregableProyecto.PHASE_CHOICES]
                        if fase not in fases_validas:
                            errores.append(f'Fila {index + 2}: Fase "{fase}" no válida. Opciones: {", ".join(fases_validas)}')
                            continue
                        
                        # Campos opcionales
                        medio = str(row.get('medio', 'Digital')).strip()
                        dossier_cliente = str(row.get('dossier_cliente', 'False')).lower() in ['true', 'sí', 'si', '1', 'yes']
                        observaciones = str(row.get('observaciones', '')).strip()
                        
                        # Fecha de entrega
                        fecha_entrega = None
                        if pd.notna(row.get('fecha_entrega')):
                            try:
                                fecha_entrega = pd.to_datetime(row['fecha_entrega']).date()
                            except:
                                # Fecha por defecto 30 días después
                                fecha_entrega = timezone.now().date() + timedelta(days=30)
                        else:
                            fecha_entrega = timezone.now().date() + timedelta(days=30)
                        
                        # Verificar si existe
                        entregable_existente = EntregableProyecto.objects.filter(
                            proyecto=proyecto, 
                            codigo=codigo
                        ).first()
                        
                        if entregable_existente:
                            if reemplazar_existentes:
                                # Actualizar existente
                                entregable_existente.nombre = nombre
                                entregable_existente.fase = fase
                                entregable_existente.creador = creador
                                entregable_existente.consolidador = consolidador
                                entregable_existente.medio = medio
                                entregable_existente.dossier_cliente = dossier_cliente
                                entregable_existente.observaciones = observaciones
                                if not entregable_existente.fecha_entrega:
                                    entregable_existente.fecha_entrega = fecha_entrega
                                entregable_existente.save()
                                entregables_actualizados += 1
                            else:
                                errores.append(f'Fila {index + 2}: Entregable con código "{codigo}" ya existe')
                                continue
                        else:
                            # Crear nuevo
                            EntregableProyecto.objects.create(
                                proyecto=proyecto,
                                codigo=codigo,
                                nombre=nombre,
                                fase=fase,
                                creador=creador,
                                consolidador=consolidador,
                                medio=medio,
                                dossier_cliente=dossier_cliente,
                                observaciones=observaciones,
                                fecha_entrega=fecha_entrega,
                                obligatorio=False,
                                seleccionado=True,
                                estado='pendiente'
                            )
                            entregables_creados += 1
                            
                    except Exception as e:
                        errores.append(f'Fila {index + 2}: Error procesando datos - {str(e)}')
                
                # Mostrar resultados
                if entregables_creados > 0 or entregables_actualizados > 0:
                    mensaje = f'Importación exitosa: {entregables_creados} entregables creados'
                    if entregables_actualizados > 0:
                        mensaje += f', {entregables_actualizados} actualizados'
                    messages.success(request, mensaje)
                
                if errores:
                    for error in errores[:5]:  # Mostrar solo los primeros 5 errores
                        messages.warning(request, error)
                    if len(errores) > 5:
                        messages.warning(request, f'... y {len(errores) - 5} errores adicionales.')
                
                if entregables_creados > 0 or entregables_actualizados > 0:
                    return redirect(f'{reverse_lazy("proyectos:gestion_entregables")}?proyecto={proyecto.id}')
                
            except Exception as e:
                messages.error(request, f'Error al procesar el archivo Excel: {str(e)}')
        
        return self.render_to_response({'form': form})


class EntregablePlantillaExcelView(LoginRequiredMixin, TemplateView):
    """Vista para generar y descargar plantilla de Excel para entregables"""
    
    def get(self, request, *args, **kwargs):
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Entregables"
        
        # Headers
        headers = [
            'codigo', 'nombre', 'fase', 'creador', 'consolidador', 
            'medio', 'dossier_cliente', 'fecha_entrega', 'observaciones'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Datos de ejemplo
        datos_ejemplo = [
            ['CUSTOM-1', 'Análisis de Requerimientos Especiales', 'Definición', 'Director de Proyecto', 'Director de Proyecto', 'Digital', 'False', '2025-07-01', 'Análisis específico para el cliente'],
            ['CUSTOM-2', 'Protocolo de Pruebas Personalizadas', 'Ejecución', 'Ing. Residente', 'Director de Proyecto', 'Digital', 'True', '2025-08-15', 'Protocolo adaptado al proyecto'],
            ['CUSTOM-3', 'Manual de Usuario Personalizado', 'Entrega', 'Director de Proyecto', 'Director de Proyecto', 'Digital', 'True', '2025-09-01', 'Manual específico del sistema instalado']
        ]
        
        for row_idx, datos in enumerate(datos_ejemplo, 2):
            for col_idx, valor in enumerate(datos, 1):
                ws.cell(row=row_idx, column=col_idx, value=valor)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Crear segunda hoja con información
        ws_info = wb.create_sheet(title="Información")
        info_data = [
            ['Campo', 'Descripción', 'Valores Válidos'],
            ['codigo', 'Código único del entregable', 'Texto único (ej: CUSTOM-1, A.1)'],
            ['nombre', 'Nombre descriptivo del entregable', 'Texto libre'],
            ['fase', 'Fase del proyecto', 'Definición, Planeación, Ejecución, Entrega'],
            ['creador', 'Responsable de crear', 'Texto libre'],
            ['consolidador', 'Responsable de consolidar', 'Texto libre'],
            ['medio', 'Medio de entrega', 'Digital, Físico, Digital/Físico'],
            ['dossier_cliente', 'Incluir en dossier', 'True/False'],
            ['fecha_entrega', 'Fecha estimada', 'YYYY-MM-DD'],
            ['observaciones', 'Comentarios adicionales', 'Texto libre (opcional)']
        ]
        
        for row_idx, info in enumerate(info_data, 1):
            for col_idx, valor in enumerate(info, 1):
                cell = ws_info.cell(row=row_idx, column=col_idx, value=valor)
                if row_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Ajustar ancho de columnas en la hoja de información
        for column in ws_info.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws_info.column_dimensions[column_letter].width = adjusted_width
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_entregables.xlsx"'
        
        # Guardar en memoria
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        
        return response
