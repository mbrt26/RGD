from django.views.generic import ListView, CreateView, UpdateView, DetailView, DeleteView, FormView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Q
from proyectos.models import Colaborador, Bitacora
from proyectos.forms.import_forms import ColaboradorImportForm
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class ColaboradorListView(LoginRequiredMixin, ListView):
    model = Colaborador
    template_name = 'proyectos/colaborador/list.html'
    context_object_name = 'colaboradores'
    paginate_by = 15
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Colaboradores'
        
        # Add filter parameters to context for maintaining state
        context['search_query'] = self.request.GET.get('search', '')
        context['cargo_filter'] = self.request.GET.get('cargo', '')
        context['has_email_filter'] = self.request.GET.get('has_email', '')
        context['has_phone_filter'] = self.request.GET.get('has_phone', '')
        
        # Get unique cargos for filter dropdown
        context['unique_cargos'] = Colaborador.objects.values_list('cargo', flat=True).distinct().order_by('cargo')
        
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Search functionality
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(nombre__icontains=search_query) |
                Q(cargo__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(telefono__icontains=search_query)
            )
        
        # Filter by cargo
        cargo_filter = self.request.GET.get('cargo')
        if cargo_filter:
            queryset = queryset.filter(cargo__icontains=cargo_filter)
        
        # Filter by email presence
        has_email = self.request.GET.get('has_email')
        if has_email == 'yes':
            queryset = queryset.exclude(email='')
        elif has_email == 'no':
            queryset = queryset.filter(email='')
        
        # Filter by phone presence
        has_phone = self.request.GET.get('has_phone')
        if has_phone == 'yes':
            queryset = queryset.exclude(telefono='')
        elif has_phone == 'no':
            queryset = queryset.filter(telefono='')
        
        return queryset.order_by('nombre')

class ColaboradorCreateView(LoginRequiredMixin, CreateView):
    model = Colaborador
    template_name = 'proyectos/colaborador/form.html'
    fields = '__all__'
    success_url = reverse_lazy('proyectos:colaborador_list')

class ColaboradorDetailView(LoginRequiredMixin, DetailView):
    model = Colaborador
    template_name = 'proyectos/colaborador/detail.html'
    context_object_name = 'colaborador'
    pk_url_kwarg = 'id'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset=queryset)
        print(f"DEBUG - Colaborador encontrado: {obj.id} - {obj.nombre}")
        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        colaborador = self.get_object()
        
        # Debug information
        print(f"DEBUG - Context data for colaborador: {colaborador.id} - {colaborador.nombre}")
        print(f"DEBUG - Email: {colaborador.email}, Teléfono: {colaborador.telefono}")
        
        # Get related bitacoras
        bitacoras = Bitacora.objects.filter(responsable=colaborador)
        print(f"DEBUG - Bitácoras encontradas: {bitacoras.count()}")
        
        context['bitacoras'] = bitacoras
        context['debug_info'] = {
            'colaborador_id': colaborador.id,
            'bitacoras_count': bitacoras.count(),
            'all_fields': {field.name: getattr(colaborador, field.name) for field in colaborador._meta.fields}
        }
        return context

class ColaboradorUpdateView(LoginRequiredMixin, UpdateView):
    model = Colaborador
    template_name = 'proyectos/colaborador/form.html'
    fields = '__all__'
    pk_url_kwarg = 'id'
    success_url = reverse_lazy('proyectos:colaborador_list')

class ColaboradorDeleteView(LoginRequiredMixin, DeleteView):
    model = Colaborador
    template_name = 'proyectos/colaborador/confirm_delete.html'
    context_object_name = 'colaborador'
    pk_url_kwarg = 'id'
    success_url = reverse_lazy('proyectos:colaborador_list')
    
    def delete(self, request, *args, **kwargs):
        colaborador = self.get_object()
        messages.success(request, f'Colaborador "{colaborador.nombre}" eliminado exitosamente.')
        return super().delete(request, *args, **kwargs)

class ColaboradorImportView(LoginRequiredMixin, FormView):
    template_name = 'proyectos/colaborador/import.html'
    form_class = ColaboradorImportForm
    success_url = reverse_lazy('proyectos:colaborador_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Importar Colaboradores'
        return context
    
    def form_valid(self, form):
        # Prevenir envío duplicado usando token de sesión
        submission_token = self.request.POST.get('submission_token')
        if submission_token and submission_token == self.request.session.get('last_submission_token'):
            messages.warning(self.request, 'Esta importación ya fue procesada.')
            return super().form_valid(form)
        
        archivo_excel = form.cleaned_data['archivo_excel']
        
        try:
            # Leer el archivo Excel
            df = pd.read_excel(archivo_excel, engine='openpyxl')
            
            # Validar columnas requeridas
            columnas_requeridas = ['nombre', 'cargo']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            
            if columnas_faltantes:
                messages.error(
                    self.request,
                    f'Error: Faltan las siguientes columnas en el archivo: {", ".join(columnas_faltantes)}'
                )
                return self.form_invalid(form)
            
            # Procesar los datos
            colaboradores_creados = 0
            colaboradores_actualizados = 0
            errores = []
            
            for index, row in df.iterrows():
                try:
                    # Obtener datos de la fila
                    nombre = str(row['nombre']).strip() if pd.notna(row['nombre']) else ''
                    cargo = str(row['cargo']).strip() if pd.notna(row['cargo']) else ''
                    email = str(row['email']).strip() if 'email' in df.columns and pd.notna(row['email']) else ''
                    telefono = str(row['telefono']).strip() if 'telefono' in df.columns and pd.notna(row['telefono']) else ''
                    
                    # Validar datos requeridos
                    if not nombre:
                        errores.append(f'Fila {index + 2}: El nombre es obligatorio')
                        continue
                    
                    if not cargo:
                        errores.append(f'Fila {index + 2}: El cargo es obligatorio')
                        continue
                    
                    # Verificar si ya existe el colaborador (por nombre y cargo)
                    colaborador_existente = Colaborador.objects.filter(
                        nombre=nombre, cargo=cargo
                    ).first()
                    
                    if colaborador_existente:
                        # Actualizar colaborador existente
                        colaborador_existente.email = email
                        colaborador_existente.telefono = telefono
                        colaborador_existente.save()
                        colaboradores_actualizados += 1
                    else:
                        # Crear nuevo colaborador
                        Colaborador.objects.create(
                            nombre=nombre,
                            cargo=cargo,
                            email=email,
                            telefono=telefono
                        )
                        colaboradores_creados += 1
                        
                except Exception as e:
                    errores.append(f'Fila {index + 2}: Error al procesar - {str(e)}')
            
            # Guardar token para prevenir duplicados
            if submission_token:
                self.request.session['last_submission_token'] = submission_token
            
            # Generar mensajes de resultado
            if colaboradores_creados > 0 or colaboradores_actualizados > 0:
                mensaje_exito = f'Importación completada: {colaboradores_creados} colaboradores creados, {colaboradores_actualizados} actualizados'
                messages.success(self.request, mensaje_exito)
            
            if errores:
                mensaje_errores = f'Se encontraron {len(errores)} errores:<br>' + '<br>'.join(errores[:10])
                if len(errores) > 10:
                    mensaje_errores += f'<br>... y {len(errores) - 10} errores más'
                messages.warning(self.request, mensaje_errores)
            
            if colaboradores_creados == 0 and colaboradores_actualizados == 0 and not errores:
                messages.info(self.request, 'No se encontraron datos válidos para importar.')
                
        except Exception as e:
            messages.error(self.request, f'Error al procesar el archivo: {str(e)}')
            return self.form_invalid(form)
        
        return super().form_valid(form)

class ColaboradorPlantillaExcelView(LoginRequiredMixin, View):
    def get(self, request):
        # Crear el workbook
        wb = Workbook()
        
        # Hoja de datos
        ws_datos = wb.active
        ws_datos.title = "Colaboradores"
        
        # Encabezados
        headers = ['nombre', 'cargo', 'email', 'telefono']
        header_descriptions = [
            'Nombre completo del colaborador (OBLIGATORIO)',
            'Cargo o posición del colaborador (OBLIGATORIO)',
            'Correo electrónico (opcional)',
            'Número de teléfono (opcional)'
        ]
        
        # Configurar encabezados
        for col, (header, description) in enumerate(zip(headers, header_descriptions), 1):
            cell = ws_datos.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
            # Agregar descripción en la segunda fila
            desc_cell = ws_datos.cell(row=2, column=col, value=description)
            desc_cell.font = Font(italic=True, size=9)
            desc_cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        # Datos de ejemplo
        ejemplos = [
            ["Juan Pérez", "Ingeniero Senior", "juan.perez@empresa.com", "300-123-4567"],
            ["María García", "Arquitecta", "maria.garcia@empresa.com", "300-234-5678"],
            ["Carlos López", "Técnico", "carlos.lopez@empresa.com", "300-345-6789"],
        ]
        
        for row_num, ejemplo in enumerate(ejemplos, 3):
            for col_num, valor in enumerate(ejemplo, 1):
                ws_datos.cell(row=row_num, column=col_num, value=valor)
        
        # Ajustar ancho de columnas
        column_widths = [25, 20, 30, 15]
        for col, width in enumerate(column_widths, 1):
            ws_datos.column_dimensions[ws_datos.cell(row=1, column=col).column_letter].width = width
        
        # Hoja de instrucciones
        ws_info = wb.create_sheet(title="Instrucciones")
        
        instrucciones = [
            "INSTRUCCIONES PARA IMPORTAR COLABORADORES",
            "",
            "1. CAMPOS OBLIGATORIOS:",
            "   - nombre: Nombre completo del colaborador",
            "   - cargo: Cargo o posición del colaborador",
            "",
            "2. CAMPOS OPCIONALES:",
            "   - email: Correo electrónico del colaborador",
            "   - telefono: Número de teléfono del colaborador",
            "",
            "3. FORMATO:",
            "   - Use la hoja 'Colaboradores' para ingresar los datos",
            "   - No modifique los nombres de las columnas",
            "   - Mantenga las columnas en el orden indicado",
            "",
            "4. NOTAS IMPORTANTES:",
            "   - Si ya existe un colaborador con el mismo nombre y cargo, se actualizará",
            "   - Los campos vacíos en columnas opcionales se ignorarán",
            "   - Elimine las filas de ejemplo antes de importar sus datos",
            "",
            "5. EJEMPLO DE USO:",
            "   - Complete los datos en la hoja 'Colaboradores'",
            "   - Guarde el archivo como .xlsx",
            "   - Use la función 'Importar Colaboradores' del sistema"
        ]
        
        for row, texto in enumerate(instrucciones, 1):
            cell = ws_info.cell(row=row, column=1, value=texto)
            if texto.startswith("INSTRUCCIONES") or texto.endswith(":"):
                cell.font = Font(bold=True)
        
        # Ajustar ancho de la columna de instrucciones
        ws_info.column_dimensions['A'].width = 70
        
        # Preparar respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_colaboradores.xlsx"'
        
        return response